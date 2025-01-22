#!/usr/bin/env python
# coding: utf-8

# In[1]:


#IndicatorCode	IndicatorName
#Ц01	Доля магистров и аспирантов в общем контингенте 
#Ц02	Средний балл ЕГЭ
#Ц03	Доля сторонних магистров и аспирантов
#Ц04	Доля целевиков
#Ц05	Объем ПОУ на ставку НПР
#Ц06	Доля остепененных ППС
#Ц07	Трудоустройство
#Ц08	Количество публикаций в WoS и Scopus и журналов из перечня ВАК на ставку НПР
#Ц09	Объем НИОКР на ставку НПР
#Ц10	Защиты аспирантов в срок
#Ц11	Доля иностранных студентов
#Ц12	Количество иностранных преподавателей
#Ц13	Академическая мобильность студентов и аспирантов
#Ц16	Заявочная активность
#Ц17	Подготовленные к публикации тематические материалы  научно-просветительского характера для неспециализированных СМИ
#Ц18	Подготовка энциклопедии/справочника
#Ц20	Руководство работами студентов и аспирантов – победителей олимпиад (конкурсов, выставок)
#Ц21	Участие в научно-технических и творческих выставках (учитывается только участие с экспонатами)
##Ц22	Членство в научно-технических или учебно-методических и редакционных советах, редколлегиях
#Ц23	Членство в программных и организационных комитетах конференций и олимпиад
#Ц24	Инновации
#Ц25	Доля ППС, СЗП которых по итогам календарного года составляет 200% и более от региональной
#Ц26	Доля общей численности НПР до 39 лет к общей численности НПР (по головам)
#Ц27	РИД


# In[3]:


import pyodbc 
import pandas as pd
import numpy as np
import datetime
import matplotlib.pyplot as plt
import math
    


# In[5]:


from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


# In[7]:


import plotly.express as px
import pandas as pd


# In[9]:


def load_formatted_scale_from_file(filename):
    """
    Загружает значение scaleEgeNapr из файла и возвращает как список.

    :param filename: имя файла для загрузки
    :return: список с диапазонами и баллами
    """
    with open(filename, "r") as file:
        scale_list = json.load(file)
    return scale_list


# In[11]:


def format_and_print_scale(scale_list, scale_name="scaleEgeNapr"):
    """
    Форматирует и выводит список в виде читаемой переменной.
    
    :param scale_list: Список с диапазонами и баллами, например:
                       [[103, 80], [102, 65], [100, 50], [99, 44], [97, 39], [95, 33], [92, 27], [90, 21], [87, 16], [0, 10]]
    :param scale_name: Имя переменной для вывода.
    """
    formatted_output = f"{scale_name} = [\n"
    for i, (threshold, score) in enumerate(scale_list):
        if i == 0:
            formatted_output += f"    ({threshold}, {score}),  # Если значение больше {threshold}, то {score} баллов\n"
        elif i == len(scale_list) - 1:
            formatted_output += f"    ({threshold}, {score})        # Если значение меньше или равно {scale_list[i-1][0]}, то {score} баллов\n"
        else:
            formatted_output += f"    ({threshold}, {score}),  # Если значение в диапазоне {threshold} – {scale_list[i-1][0]}, то {score} баллов\n"
    formatted_output += "]"
    print(formatted_output)


# In[ ]:





# In[14]:


# загрузка шкал баллов за выполнение плана


# In[16]:


scaleEgeVip = load_formatted_scale_from_file("Shkali/scaleEgeVip.json")
scaleTrudoustroistvoVip = load_formatted_scale_from_file("Shkali/scaleTrudoustroistvoVip.json")
scaleRidVip = load_formatted_scale_from_file("Shkali/scaleRidVip.json")
scalePublNauchMaterialVip = load_formatted_scale_from_file("Shkali/scalePublNauchMaterialVip.json")
scaleAkadMobVip = load_formatted_scale_from_file("Shkali/scaleAkadMobVip.json")
scaleDolInostrVip = load_formatted_scale_from_file("Shkali/scaleDolInostrVip.json")
scaleDolMagVip = load_formatted_scale_from_file("Shkali/scaleDolMagVip.json")
scaleDolOstepenennykhVip = load_formatted_scale_from_file("Shkali/scaleDolOstepenennykhVip.json")
scaleDolPPS200Vip = load_formatted_scale_from_file("Shkali/scaleDolPPS200Vip.json")
scaleDolStorMagVip = load_formatted_scale_from_file("Shkali/scaleDolStorMagVip.json")
scaleDolCelevikovVip = load_formatted_scale_from_file("Shkali/scaleDolCelevikovVip.json")
scaleZashchityVip = load_formatted_scale_from_file("Shkali/scaleZashchityVip.json")
scaleZayavAktivVip = load_formatted_scale_from_file("Shkali/scaleZayavAktivVip.json")
scaleInnovaciiVip = load_formatted_scale_from_file("Shkali/scaleInnovaciiVip.json")
scaleKolInostrPrepodVip = load_formatted_scale_from_file("Shkali/scaleKolInostrPrepodVip.json")
scaleKolPublikVip = load_formatted_scale_from_file("Shkali/scaleKolPublikVip.json")
scaleObemNIOKRVip = load_formatted_scale_from_file("Shkali/scaleObemNIOKRVip.json")
scaleObemPOUVip = load_formatted_scale_from_file("Shkali/scaleObemPOUVip.json")
scaleDolMolNPRVip = load_formatted_scale_from_file("Shkali/scaleDolMolNPRVip.json")


# In[18]:


scaleEgeVip = [
    (109.32, 90.000000),  # Если значение больше 109.32%, то 80.0 баллов
    (100.79, 85.000000),  # Если значение в диапазоне 100.79% – 109.32%, то 65.0 баллов
    (98.91, 80.000000),  # Если значение в диапазоне 98.91% – 100.79%, то 50.0 баллов
    (94.49, 75.000000),  # Если значение в диапазоне 94.49% – 98.91%, то 50.0 баллов
    (91.31, 70.000000),  # Если значение в диапазоне 91.31% – 94.49%, то 50.0 баллов
    (90.09, 70.000000),  # Если значение в диапазоне 90.09% – 91.31%, то 50.0 баллов
    (88.38, 65.000000),  # Если значение в диапазоне 88.38% – 90.09%, то 50.0 баллов
    (87.02, 65.000000),  # Если значение в диапазоне 87.02% – 88.38%, то 50.0 баллов
    (85.44, 65.000000),  # Если значение в диапазоне 85.44% – 87.02%, то 50.0 баллов
    (50,    65.000000),  # Если значение в диапазоне 85.44% – 87.02%, то 50.0 баллов
    (0, 10)        # Если значение меньше или равно 85.44%, то 10 баллов
]


# In[ ]:





# In[21]:


# загрузка шкал напряженностей


# In[23]:


scaleEgeNapr = load_formatted_scale_from_file("Shkali/scaleEgeNapr.json")
scaleTrudoustroistvoNapr = load_formatted_scale_from_file("Shkali/scaleTrudoustroistvoNapr.json")
scaleRidNapr = load_formatted_scale_from_file("Shkali/scaleRidNapr.json")
scalePublNauchMaterialNapr = load_formatted_scale_from_file("Shkali/scalePublNauchMaterialNapr.json")
scaleAkadMobNapr = load_formatted_scale_from_file("Shkali/scaleAkadMobNapr.json")
scaleDolInostrNapr = load_formatted_scale_from_file("Shkali/scaleDolInostrNapr.json")
scaleDolMagNapr = load_formatted_scale_from_file("Shkali/scaleDolMagNapr.json")
scaleDolOstepenennykhNapr = load_formatted_scale_from_file("Shkali/scaleDolOstepenennykhNapr.json")
scaleDolPPS200Napr = load_formatted_scale_from_file("Shkali/scaleDolPPS200Napr.json")
scaleDolStorMagNapr = load_formatted_scale_from_file("Shkali/scaleDolStorMagNapr.json")
scaleDolCelevikovNapr = load_formatted_scale_from_file("Shkali/scaleDolCelevikovNapr.json")
scaleZashchityNapr = load_formatted_scale_from_file("Shkali/scaleZashchityNapr.json")
scaleZayavAktivNapr = load_formatted_scale_from_file("Shkali/scaleZayavAktivNapr.json")
scaleInnovaciiNapr = load_formatted_scale_from_file("Shkali/scaleInnovaciiNapr.json")
scaleKolInostrPrepodNapr = load_formatted_scale_from_file("Shkali/scaleKolInostrPrepodNapr.json")
scaleKolPublikNapr = load_formatted_scale_from_file("Shkali/scaleKolPublikNapr.json")
scaleObemNIOKRNapr = load_formatted_scale_from_file("Shkali/scaleObemNIOKRNapr.json")
scaleObemPOUNapr = load_formatted_scale_from_file("Shkali/scaleObemPOUNapr.json")
scaleDolMolNPRNapr = load_formatted_scale_from_file("Shkali/scaleDolMolNPRNapr.json")


# In[25]:


# Список переменных и их имен
scales = [
    ("scaleEgeNapr", scaleEgeNapr),
    ("scaleTrudoustroistvoNapr", scaleTrudoustroistvoNapr),
    ("scaleRidNapr", scaleRidNapr),
    ("scalePublNauchMaterialNapr", scalePublNauchMaterialNapr),
    ("scaleAkadMobNapr", scaleAkadMobNapr),
    ("scaleDolInostrNapr", scaleDolInostrNapr),
    ("scaleDolMagNapr", scaleDolMagNapr),
    ("scaleDolOstepenennykhNapr", scaleDolOstepenennykhNapr),
    ("scaleDolPPS200Napr", scaleDolPPS200Napr),
    ("scaleDolStorMagNapr", scaleDolStorMagNapr),
    ("scaleDolCelevikovNapr", scaleDolCelevikovNapr),
    ("scaleZashchityNapr", scaleZashchityNapr),
    ("scaleZayavAktivNapr", scaleZayavAktivNapr),
    ("scaleInnovaciiNapr", scaleInnovaciiNapr),
    ("scaleKolInostrPrepodNapr", scaleKolInostrPrepodNapr),
    ("scaleKolPublikNapr", scaleKolPublikNapr),
    ("scaleObemNIOKRNapr", scaleObemNIOKRNapr),
    ("scaleObemPOUNapr", scaleObemPOUNapr),
    ("scaleDolMolNPRNapr", scaleDolMolNPRNapr)
]

# Вызываем функцию для каждой шкалы
for name, scale in scales:
    print(f"Форматированная шкала для {name}:")
    format_and_print_scale(scale, scale_name=name)
    print("\n")  # Пустая строка для разделения вывода


# In[27]:


# Список переменных и их имен
scales = [
    ("scaleEgeVip", scaleEgeVip),
    ("scaleTrudoustroistvoVip", scaleTrudoustroistvoVip),
    ("scaleRidVip", scaleRidVip),
    ("scalePublNauchMaterialVip", scalePublNauchMaterialVip),
    ("scaleAkadMobVip", scaleAkadMobVip),
    ("scaleDolInostrVip", scaleDolInostrVip),
    ("scaleDolMagVip", scaleDolMagVip),
    ("scaleDolOstepenennykhVip", scaleDolOstepenennykhVip),
    ("scaleDolPPS200Vip", scaleDolPPS200Vip),
    ("scaleDolStorMagVip", scaleDolStorMagVip),
    ("scaleDolCelevikovVip", scaleDolCelevikovVip),
    ("scaleZashchityVip", scaleZashchityVip),
    ("scaleZayavAktivVip", scaleZayavAktivVip),
    ("scaleInnovaciiVip", scaleInnovaciiVip),
    ("scaleKolInostrPrepodVip", scaleKolInostrPrepodVip),
    ("scaleKolPublikVip", scaleKolPublikVip),
    ("scaleObemNIOKRVip", scaleObemNIOKRVip),
    ("scaleObemPOUVip", scaleObemPOUVip),
    ("scaleDolMolNPRVip", scaleDolMolNPRVip)
]

# Вызываем функцию для каждой шкалы
for name, scale in scales:
    print(f"Форматированная шкала для {name}:")
    format_and_print_scale(scale, scale_name=name)
    print("\n")  # Пустая строка для разделения вывода


# In[29]:


scaleEgeNapr = [
 [1.16, 90],
 [1.14, 85],
 [1.09, 80],
 [1.06, 75],
 [1.04, 70],
 [1.02, 65],
 [0.99, 60],
 [0.97, 50],
 [0.93, 50],
 [0.80, 50],
 [0.70, 50],
 [0.60, 50],
 [0, 10] 
]


# In[31]:


# Функция для расчета баллов на основе шкалы
def CalcBallVipolnenia(value, scale):
    """
    value: float - значение выполнения плана
    scale: list of tuples - шкала в формате [(процент, баллы), ...]
    """
    for percent, score in scale:
        if value >= percent:
            return score
    # Если значение меньше минимального, возвращаем минимальный балл
    return scale[-1][1]

# Пример использования
#example_value = 93  # Значение выполнения плана
#score = CalcBallVipolnenia(example_value, scaleTrud)
#print(f"Баллы за выполнение: {score}")


# In[33]:


def CalcBallCompare(E_kaf, scale, E_univ):
    
    """
    value: float - значение выполнения плана
    scale: list of tuples - шкала в формате [(процент, баллы), ...]
    """
    for percent, score in scale:
        if (E_kaf/E_univ) >= percent:
            return score
    # Если значение меньше минимального, возвращаем минимальный балл
    return scale[-1][1]


# In[35]:


def calculate_final_score(B_vyp, B_srav, w_vyp=0.55, w_srav=0.45):
    # Расчет итогового балла по формуле
    B_itog = w_vyp * B_vyp + w_srav * B_srav
    # Округление в большую сторону до числа, кратного 5
    B_itog = math.ceil(B_itog / 5) * 5
    # Применение ограничений
    if B_itog > 100:
        B_itog = 100
    elif B_itog <= 10 and B_itog>0.01:
        B_itog = 10
    elif B_itog<=0.01:
        B_itog = 0
    
    return B_itog


# In[37]:


def get_sredBallFinal_values(df, kafedra_column, sredBallVipolnenia_column, sredBallCompare_column, sredBallFinal_column):
    # Обновленный список нужных кафедр в заданном порядке
    kafedry_list = [
        "ЭЭС", "РЗиАЭ", "ИТНО", "ЭППЭ", "Пром.эл.", "МЭП", "ТЭС", "ТЭВН", 
        "РМДиПМ", "ПТС", "ЭМЭЭА", "ИЭиОТ", "УИТ", "ТМ", "Физики", "ЭЭП", 
        "ГВИЭ", "ОФиЯС", "АЭС", "ТОЭ", "АСУТП", "И и К", "ПГТ", "РТС", 
        "ЭиН", "АЭП", "ХиЭЭ", "ВТ", "МиПЭУ", "ПМИИ", "ВМСС", "Эл.ст.", 
        "НТ", "ЭКАОиЭТ", "ТОТ", "МКМ", "ГГМ", "ДИТ", "ТМПУ", "РС и Л", 
        "ОРТ", "ФТЭМК", "Ин.яз", "БИТ", "ИТФ", "РТП и АС", "ЭГТС", "ФОРС", 
        "Светотех.", "ФПС", "Дизайн", "ВМ", "Ф и С"
    ]

    # Фильтрация DataFrame по списку кафедр
    filtered_df = df[df[kafedra_column].isin(kafedry_list)]

    # Сохранение результатов в нужном порядке
    result = []
    for kafedra in kafedry_list:
        # Находим строки, соответствующие конкретной кафедре
        row = filtered_df[filtered_df[kafedra_column] == kafedra]
        if not row.empty:
            # Извлекаем значения из нужных столбцов и приводим к целому числу
            vipolnenia = int(row.iloc[0][sredBallVipolnenia_column])
            compare = int(row.iloc[0][sredBallCompare_column])
            final = int(row.iloc[0][sredBallFinal_column])
            result.append(f"{kafedra}; {vipolnenia}; {compare}; {final}")
        else:
            # Если данных по кафедре нет, выводим "-"
            result.append(f"{kafedra}; - ; - ; -")

    return result


# In[39]:


pwd


# In[49]:


# Сохраняем DataFrame в CSV файл
#df.to_csv('dannie6.csv', index=False)

# Очищаем DataFrame
df = pd.DataFrame()

# Загружаем данные из CSV файла
#df = pd.read_csv('dannie24.csv')
df = pd.read_csv('dannie22_01_25.csv')


# In[53]:


# Загрузка данных из CSV файла средних по университету
dfSred = pd.read_csv('sredneeUniver24.csv', delimiter=';')

# Преобразование DataFrame в словарь
dictSred = dfSred.set_index('IndicatorCode')['SredneePoUniver'].to_dict()

print(dictSred)


# In[55]:


df


# In[57]:


columns_list = df.columns.tolist()
print(columns_list)


# In[59]:


# Фильтрация строк, где кафедра равна 'РТС'
filtered_df = df[df['kafedra'] == 'РТС']

# Вывод отфильтрованных данных
filtered_df.to_csv('filtered_data.csv', index=False, encoding='utf-8')


# In[ ]:





# In[62]:


# Преобразование значений в словаре dictSred, заменяя запятую на точку перед преобразованием в float
dictSred = {key: float(value.replace(',', '.')) for key, value in dictSred.items()}


# In[64]:


def process_data(EGEFact, EGEVip, EGEFactVipolnenia, df, scaleEgeVip, scaleEgeNapr,
                 EGEFactBallCompare, EGEFactFinal, sred_C02):
    
      
    # Приведение значений в 'EGEFact' к float
    df[EGEFact] = pd.to_numeric(df[EGEFact], errors='coerce')
    
    # Приведение значений в 'EGEVip' к float
    df[EGEVip] = pd.to_numeric(df[EGEVip], errors='coerce')
    
    # Приведение значения sred_C02 к float с заменой запятой на точку
    if isinstance(sred_C02, str):
        sred_C02 = float(sred_C02.replace(',', '.'))
    else:
        sred_C02 = float(sred_C02)
    
    # Применяем функцию CalcBallVipolnenia для столбца EGEFactVipolnenia
    df[EGEFactVipolnenia] = df[EGEVip].apply(
        lambda x: CalcBallVipolnenia(x, scaleEgeVip) if x >= 0.002 else 0
    )
    
    # Применяем функцию CalcBallCompare для столбца EGEFactBallCompare
    df[EGEFactBallCompare] = df[EGEFact].apply(
        lambda x: CalcBallCompare(x, scaleEgeNapr,sred_C02) if x >= 0.002 else 0
    )
    
    # Применяем функцию calculate_final_score для столбца EGEFactFinal
    df[EGEFactFinal] = df.apply(
        lambda row: calculate_final_score(row[EGEFactVipolnenia], row[EGEFactBallCompare]),
        axis=1
    )
    
    return df


# In[66]:


def calculate_krit6(row, columns):
    # Извлекаем ненулевые значения из указанных столбцов
    non_zero_values = [value for value in row[columns] if value > 0]
    
    # Если ненулевых значений нет, возвращаем 0
    if not non_zero_values:
        return 0
    
    # Вычисляем среднее ненулевых значений
    avg_value = sum(non_zero_values) / len(non_zero_values)
    
    # Округляем до ближайшего кратного 5
    return round(avg_value / 5) * 5


# In[68]:


print(list(df.columns))


# In[70]:


scaleDolMagVip


# In[72]:


scaleDolMagNapr


# In[74]:


def plot_columns_by_kafedra(df, x_col, y_col_1, y_col_2=None):
    """
    Универсальная функция для построения графиков зависимости y_col_1 (и y_col_2, если задано)
    отдельно для каждой кафедры.

    Параметры:
    - df (pd.DataFrame): Датафрейм с данными.
    - x_col (str): Название столбца для оси X.
    - y_col_1 (str): Название первого столбца для оси Y.
    - y_col_2 (str, optional): Название второго столбца для оси Y.
    """
    kafedras = df[x_col].unique()  # Уникальные кафедры
    
    for kafedra in kafedras:
        # Отфильтровать данные для текущей кафедры
        kafedra_data = df[df[x_col] == kafedra]
        
        plt.figure(figsize=(8, 5))
        
        # График для y_col_1
        plt.plot(kafedra_data.index, kafedra_data[y_col_1], marker='o', label=y_col_1)
        
        # График для y_col_2 (если передан)
        if y_col_2:
            plt.plot(kafedra_data.index, kafedra_data[y_col_2], marker='o', label=y_col_2)
        
        # Настройки графика
        plt.xlabel("Index")
        plt.ylabel("Values")
        plt.title(f"Graph for {kafedra}: {y_col_1} and {y_col_2} (if provided)")
        plt.legend()
        plt.grid()
        plt.tight_layout()
        plt.show()


# In[76]:


def plot_scrollable_graph(df, x_col, y_col_1, y_col_2=None):
    """
    Создает прокручиваемый график зависимости y_col_1 (и y_col_2, если задано) для всех кафедр.

    Параметры:
    - df (pd.DataFrame): Датафрейм с данными.
    - x_col (str): Название столбца для оси X (например, кафедры).
    - y_col_1 (str): Название первого столбца для оси Y.
    - y_col_2 (str, optional): Название второго столбца для оси Y.
    """
    # Создаем длинный DataFrame для удобного отображения в Plotly
    data = df.melt(id_vars=[x_col], value_vars=[y_col_1, y_col_2] if y_col_2 else [y_col_1],
                   var_name='Metric', value_name='Value')

    # Строим интерактивный график
    fig = px.line(
        data,
        x=x_col,
        y="Value",
        color="Metric",
        title=f"Scrollable Graph: {y_col_1} and {y_col_2 if y_col_2 else ''}",
        labels={x_col: "Kafedra", "Value": "Values", "Metric": "Metrics"},
    )

    # Настройка оси X для прокрутки
    fig.update_xaxes(rangeslider_visible=True)

    # Отображаем график
    fig.show()


# In[78]:


def save_df_with_scales_to_excel(df, columns, filename, sheet_name, scale_vip, scale_napr):
    """
    Сохраняет указанные столбцы DataFrame в Excel с добавлением справа значений scaleDolMagVip и scaleDolMagNapr.
    Столбцы для "Баллы за выполнение плана" и "Баллы за напряженность" сдвигаются на 6 столбцов.
    Автоподбор ширины столбцов добавлен.

    Параметры:
    - df (pd.DataFrame): Датафрейм для записи.
    - columns (list): Список столбцов для сохранения.
    - filename (str): Имя выходного файла.
    - sheet_name (str): Название листа.
    - scale_vip (list): Список значений для scaleDolMagVip.
    - scale_napr (list): Список значений для scaleDolMagNapr.
    """
    # Выбираем только указанные столбцы
    selected_df = df[columns]

    # Создаем новый Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Добавляем DataFrame в лист
    for r_idx, row in enumerate(dataframe_to_rows(selected_df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Сдвиг на 6 столбцов для scaleDolMagVip и scaleDolMagNapr
    vip_start_col = ws.max_column + 6  # Колонка для размещения scaleDolMagVip
    ws.cell(row=1, column=vip_start_col, value="Баллы за выполнение плана")
    for i, value in enumerate(scale_vip, start=2):
        ws.cell(row=i, column=vip_start_col, value=str(value))

    napr_start_col = vip_start_col + 1  # Колонка для размещения scaleDolMagNapr
    ws.cell(row=1, column=napr_start_col, value="Баллы за напряженность")
    for i, value in enumerate(scale_napr, start=2):
        ws.cell(row=i, column=napr_start_col, value=str(value))

    # Устанавливаем границы вокруг всех данных
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Автоподбор ширины столбцов
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Получаем букву столбца
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Добавляем запас ширины

    # Сохраняем файл
    wb.save(filename)



# In[80]:


#По всем показателям


# In[82]:


# Для Доля магистров и аспирантов в общем контингенте
result_df = process_data(
    EGEFact='DolMagFact',
    EGEVip='DolMagVip',
    EGEFactVipolnenia='DolMagFactVipolnenia',
    df=df,
    scaleEgeVip=scaleDolMagVip,
    scaleEgeNapr = scaleDolMagNapr,
    EGEFactBallCompare='DolMagFactBallCompare',
    EGEFactFinal='DolMagFactFinal',
    sred_C02=dictSred['Ц01']
)
result_df['DolMagNapr'] = result_df['DolMagFact'] / dictSred['Ц01']
result_df[['kafedra', 'DolMagFact', 'DolMagVip', 'DolMagNapr', 'DolMagFactVipolnenia', 'DolMagFactBallCompare', 'DolMagFactFinal']]


# In[84]:


scaleDolMagVip


# In[86]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'DolMagFact', 'DolMagVip', 'DolMagNapr', 'DolMagFactVipolnenia', 'DolMagFactBallCompare', 'DolMagFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Доля магистров и аспирантов в общем контингенте.xlsx', 'Доля магистров', scaleDolMagVip, scaleDolMagNapr)


# In[ ]:





# In[ ]:





# In[90]:


plot_scrollable_graph(result_df, 'kafedra', 'DolMagFactVipolnenia', 'DolMagVip')


# In[92]:


plot_scrollable_graph(result_df, 'kafedra', 'DolMagFact', 'DolMagFactBallCompare')


# In[94]:


scaleEgeVip


# In[96]:


scaleEgeNapr


# In[98]:


# Для Средний балл ЕГЭ
result_df = process_data(
    EGEFact='EGEFact',
    EGEVip='EGEVip',
    EGEFactVipolnenia='EGEFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleEgeVip,
    scaleEgeNapr = scaleEgeNapr,
    EGEFactBallCompare='EGEFactBallCompare',
    EGEFactFinal='EGEFactFinal',
    sred_C02=dictSred['Ц02']
)
result_df['EGENapr'] = result_df['EGEFact'] / dictSred['Ц02']
result_df[['kafedra', 'EGEFact', 'EGEVip', 'EGENapr','EGEFactVipolnenia', 'EGEFactBallCompare', 'EGEFactFinal']]


# In[100]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'EGEFact', 'EGEVip', 'EGENapr','EGEFactVipolnenia', 'EGEFactBallCompare', 'EGEFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'ЕГЭ.xlsx', 'ЕГЭ', scaleEgeVip, scaleEgeNapr)


# In[ ]:





# In[ ]:





# In[104]:


# Для Доля сторонних магистров и аспирантов
result_df = process_data(
    EGEFact='DolStorMagFact',
    EGEVip='DolStorMagVip',
    EGEFactVipolnenia='DolStorMagFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleDolStorMagVip,
    scaleEgeNapr = scaleDolStorMagNapr,
    EGEFactBallCompare='DolStorMagFactBallCompare',
    EGEFactFinal='DolStorMagFactFinal',
    sred_C02=dictSred['Ц03']
)


# In[106]:


result_df['DolStorMagNapr'] = result_df['DolStorMagFact'] / dictSred['Ц03']
result_df[['kafedra', 'DolStorMagFact', 'DolStorMagVip', 'DolStorMagNapr', 'DolStorMagFactVipolnenia', 'DolStorMagFactBallCompare', 'DolStorMagFactFinal']]


# In[108]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'DolStorMagFact', 'DolStorMagVip', 'DolStorMagNapr', 'DolStorMagFactVipolnenia', 'DolStorMagFactBallCompare', 'DolStorMagFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Доля сторонних магистров и аспирантов.xlsx', 'Доля Доля сторонних магистров', scaleDolStorMagVip, scaleDolStorMagNapr)


# In[110]:


# Для Доля целевиков
result_df = process_data(
    EGEFact='DolCelevikovFact',
    EGEVip='DolCelevikovVip',
    EGEFactVipolnenia='DolCelevikovFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleDolCelevikovVip,
    scaleEgeNapr = scaleDolCelevikovNapr,
    EGEFactBallCompare='DolCelevikovFactBallCompare',
    EGEFactFinal='DolCelevikovFactFinal',
    sred_C02=dictSred['Ц04']
)
result_df['DolCelevikovNapr'] = result_df['DolCelevikovFact'] / dictSred['Ц04']
result_df[['kafedra', 'DolCelevikovFact', 'DolCelevikovVip', 'DolCelevikovNapr','DolCelevikovFactVipolnenia', 'DolCelevikovFactBallCompare', 'DolCelevikovFactFinal']]


# In[112]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'DolCelevikovFact', 'DolCelevikovVip', 'DolCelevikovNapr','DolCelevikovFactVipolnenia', 'DolCelevikovFactBallCompare', 'DolCelevikovFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Доля целевиков.xlsx', 'Доля целевиков', scaleDolCelevikovVip, scaleDolCelevikovNapr)


# In[114]:


# Для Объем ПОУ на ставку НПР
result_df = process_data(
    EGEFact='ObemPOUFact',
    EGEVip='ObemPOUVip',
    EGEFactVipolnenia='ObemPOUFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleObemPOUVip,
    scaleEgeNapr = scaleObemPOUNapr,
    EGEFactBallCompare='ObemPOUFactBallCompare',
    EGEFactFinal='ObemPOUFactFinal',
    sred_C02=dictSred['Ц05']
)
result_df['ObemPOUNapr'] = result_df['ObemPOUFact'] / dictSred['Ц05']
result_df[['kafedra', 'ObemPOUFact', 'ObemPOUVip', 'ObemPOUNapr', 'ObemPOUFactVipolnenia', 'ObemPOUFactBallCompare', 'ObemPOUFactFinal']]


# In[116]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'ObemPOUFact', 'ObemPOUVip', 'ObemPOUNapr', 'ObemPOUFactVipolnenia', 'ObemPOUFactBallCompare', 'ObemPOUFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Объем ПОУ на ставку НПР.xlsx', 'Объем ПОУ', scaleObemPOUVip, scaleObemPOUNapr)


# In[ ]:





# In[119]:


# Для Доля остепененных ППС
result_df = process_data(
    EGEFact='DolOstepenennykhFact',
    EGEVip='DolOstepenennykhVip',
    EGEFactVipolnenia='DolOstepenennykhFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleDolOstepenennykhVip,
    scaleEgeNapr = scaleDolOstepenennykhNapr,
    EGEFactBallCompare='DolOstepenennykhFactBallCompare',
    EGEFactFinal='DolOstepenennykhFactFinal',
    sred_C02=dictSred['Ц06']
)
result_df['DolOstepenennykhNapr'] = result_df['DolOstepenennykhFact'] / dictSred['Ц06']
result_df[['kafedra', 'DolOstepenennykhFact', 'DolOstepenennykhVip', 'DolOstepenennykhNapr','DolOstepenennykhFactVipolnenia', 'DolOstepenennykhFactBallCompare', 'DolOstepenennykhFactFinal']]


# In[121]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'DolOstepenennykhFact', 'DolOstepenennykhVip', 'DolOstepenennykhNapr','DolOstepenennykhFactVipolnenia', 'DolOstepenennykhFactBallCompare', 'DolOstepenennykhFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Доля остепененных ППС.xlsx', 'Доля остепененных ППС', scaleDolOstepenennykhVip, scaleDolOstepenennykhNapr)


# In[123]:


# Для Трудоустройства
result_df = process_data(
    EGEFact='TrudoustroistvoFact',
    EGEVip='TrudoustroistvoVip',
    EGEFactVipolnenia='TrudoustroistvoFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleTrudoustroistvoVip,
    scaleEgeNapr = scaleTrudoustroistvoNapr,
    EGEFactBallCompare='TrudoustroistvoFactBallCompare',
    EGEFactFinal='TrudoustroistvoFactFinal',
    sred_C02=dictSred['Ц07']
)
result_df['TrudoustroistvoNapr'] = result_df['TrudoustroistvoFact'] / dictSred['Ц07']
result_df[['kafedra', 'TrudoustroistvoFact', 'TrudoustroistvoVip', 'TrudoustroistvoNapr', 'TrudoustroistvoFactVipolnenia', 'TrudoustroistvoFactBallCompare', 'TrudoustroistvoFactFinal']]


# In[125]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'TrudoustroistvoFact', 'TrudoustroistvoVip', 'TrudoustroistvoNapr', 'TrudoustroistvoFactVipolnenia', 'TrudoustroistvoFactBallCompare', 'TrudoustroistvoFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Трудоустройство.xlsx', 'Трудоустройство', scaleTrudoustroistvoVip, scaleTrudoustroistvoNapr)


# In[ ]:





# In[128]:


# Для Количество публикаций
result_df = process_data(
    EGEFact='KolPublikFact',
    EGEVip='KolPublikVip',
    EGEFactVipolnenia='KolPublikFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleKolPublikVip,
    scaleEgeNapr = scaleKolPublikNapr,
    EGEFactBallCompare='KolPublikFactBallCompare',
    EGEFactFinal='KolPublikFactFinal',
    sred_C02=dictSred['Ц08']
)
result_df['KolPublikNapr'] = result_df['KolPublikFact'] / dictSred['Ц08']
result_df[['kafedra', 'KolPublikFact', 'KolPublikVip', 'KolPublikNapr', 'KolPublikFactVipolnenia', 'KolPublikFactBallCompare', 'KolPublikFactFinal']]


# In[130]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'KolPublikFact', 'KolPublikVip', 'KolPublikNapr', 'KolPublikFactVipolnenia', 'KolPublikFactBallCompare', 'KolPublikFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Количество публикаций в WoS и Scopus и журналов из перечня ВАК на ставку НПР.xlsx', 'Количество публикаций в WoS', scaleKolPublikVip, scaleKolPublikNapr)


# In[132]:


# Для Объем НИОКР
result_df = process_data(
    EGEFact='ObemNIOKRFact',
    EGEVip='ObemNIOKRVip',
    EGEFactVipolnenia='ObemNIOKRFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleObemNIOKRVip,
    scaleEgeNapr = scaleObemNIOKRNapr,
    EGEFactBallCompare='ObemNIOKRFactBallCompare',
    EGEFactFinal='ObemNIOKRFactFinal',
    sred_C02=dictSred['Ц09']
)
result_df['ObemNIOKRNapr'] = result_df['ObemNIOKRFact'] / dictSred['Ц09']
result_df[['kafedra', 'ObemNIOKRFact', 'ObemNIOKRVip', 'ObemNIOKRNapr', 'ObemNIOKRFactVipolnenia', 'ObemNIOKRFactBallCompare', 'ObemNIOKRFactFinal']]


# In[134]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'ObemNIOKRFact', 'ObemNIOKRVip', 'ObemNIOKRNapr', 'ObemNIOKRFactVipolnenia', 'ObemNIOKRFactBallCompare', 'ObemNIOKRFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Объем НИОКР на ставку НПР.xlsx', 'Объем НИОКР на ставку НПР', scaleObemNIOKRVip, scaleObemNIOKRNapr)


# In[136]:


# Для Защиты аспирантов в срок
result_df = process_data(
    EGEFact='ZashchityFact',
    EGEVip='ZashchityVip',
    EGEFactVipolnenia='ZashchityFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleZashchityVip,
    scaleEgeNapr = scaleZashchityNapr,
    EGEFactBallCompare='ZashchityFactBallCompare',
    EGEFactFinal='ZashchityFactFinal',
    sred_C02=dictSred['Ц10']
)
result_df['ZashchityNapr'] = result_df['ZashchityFact'] / dictSred['Ц10']
result_df[['kafedra', 'ZashchityFact', 'ZashchityVip', 'ZashchityNapr', 'ZashchityFactVipolnenia', 'ZashchityFactBallCompare', 'ZashchityFactFinal']]


# In[138]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'ZashchityFact', 'ZashchityVip', 'ZashchityNapr', 'ZashchityFactVipolnenia', 'ZashchityFactBallCompare', 'ZashchityFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Защиты аспирантов в срок.xlsx', 'Защиты аспирантов в срок', scaleZashchityVip, scaleZashchityNapr)


# In[ ]:





# In[141]:


# Для Доля иностранных студентов
result_df = process_data(
    EGEFact='DolInostrFact',
    EGEVip='DolInostrVip',
    EGEFactVipolnenia='DolInostrFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleDolInostrVip,
    scaleEgeNapr = scaleDolInostrNapr,
    EGEFactBallCompare='DolInostrFactBallCompare',
    EGEFactFinal='DolInostrFactFinal',
    sred_C02=dictSred['Ц11']
)
result_df['DolInostrNapr'] = result_df['DolInostrFact'] / dictSred['Ц11']
result_df[['kafedra', 'DolInostrFact', 'DolInostrVip', 'DolInostrNapr','DolInostrFactVipolnenia', 'DolInostrFactBallCompare', 'DolInostrFactFinal']]


# In[143]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'DolInostrFact', 'DolInostrVip', 'DolInostrNapr','DolInostrFactVipolnenia', 'DolInostrFactBallCompare', 'DolInostrFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Доля иностранных студентов.xlsx', 'Доля иностранных студентов', scaleDolInostrVip, scaleDolInostrNapr)


# In[145]:


# Для Количество иностранных преподавателей
result_df = process_data(
    EGEFact='KolInostrPrepodFact',
    EGEVip='KolInostrPrepodVip',
    EGEFactVipolnenia='KolInostrPrepodFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleKolInostrPrepodVip,
    scaleEgeNapr = scaleKolInostrPrepodNapr,
    EGEFactBallCompare='KolInostrPrepodFactBallCompare',
    EGEFactFinal='KolInostrPrepodFactFinal',
    sred_C02=dictSred['Ц12']
)
result_df['KolInostrPrepodNapr'] = result_df['KolInostrPrepodFact'] / dictSred['Ц12']
result_df[['kafedra', 'KolInostrPrepodFact', 'KolInostrPrepodVip', 'KolInostrPrepodNapr', 'KolInostrPrepodFactVipolnenia', 'KolInostrPrepodFactBallCompare', 'KolInostrPrepodFactFinal']]


# In[147]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'KolInostrPrepodFact', 'KolInostrPrepodVip', 'KolInostrPrepodNapr', 'KolInostrPrepodFactVipolnenia', 'KolInostrPrepodFactBallCompare', 'KolInostrPrepodFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Количество иностранных преподавателей.xlsx', 'Количество иностранных преподавателей', scaleKolInostrPrepodVip, scaleKolInostrPrepodNapr)


# In[149]:


# Для Академическая мобильность студентов и аспирантов
result_df = process_data(
    EGEFact='AkadMobFact',
    EGEVip='AkadMobVip',
    EGEFactVipolnenia='AkadMobFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleAkadMobVip,
    scaleEgeNapr = scaleAkadMobNapr,
    EGEFactBallCompare='AkadMobFactBallCompare',
    EGEFactFinal='AkadMobFactFinal',
    sred_C02=dictSred['Ц13']
)
result_df['AkadMobNapr'] = result_df['AkadMobFact'] / dictSred['Ц13']
result_df[['kafedra', 'AkadMobFact', 'AkadMobVip', 'AkadMobNapr','AkadMobFactVipolnenia', 'AkadMobFactBallCompare', 'AkadMobFactFinal']]


# In[151]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'AkadMobFact', 'AkadMobVip', 'AkadMobNapr','AkadMobFactVipolnenia', 'AkadMobFactBallCompare', 'AkadMobFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Академическая мобильность.xlsx', 'Академическая мобильность', scaleAkadMobVip, scaleAkadMobNapr)


# In[153]:


# Для Заявочная активность
result_df = process_data(
    EGEFact='ZayavAktivFact',
    EGEVip='ZayavAktivVip',
    EGEFactVipolnenia='ZayavAktivFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleZayavAktivVip,
    scaleEgeNapr = scaleZayavAktivNapr,
    EGEFactBallCompare='ZayavAktivFactBallCompare',
    EGEFactFinal='ZayavAktivFactFinal',
    sred_C02=dictSred['Ц16']
)
result_df['ZayavAktivNapr'] = result_df['ZayavAktivFact'] / dictSred['Ц16']
result_df[['kafedra', 'ZayavAktivFact', 'ZayavAktivVip', 'ZayavAktivNapr', 'ZayavAktivFactVipolnenia', 'ZayavAktivFactBallCompare', 'ZayavAktivFactFinal']]


# In[155]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'ZayavAktivFact', 'ZayavAktivVip', 'ZayavAktivNapr', 'ZayavAktivFactVipolnenia', 'ZayavAktivFactBallCompare', 'ZayavAktivFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Заявочная активность.xlsx', 'Заявочная активность', scaleZayavAktivVip, scaleZayavAktivNapr)


# In[ ]:





# In[ ]:





# In[159]:


# Для Подготовленные к публикации тематические материалы
result_df = process_data(
    EGEFact='PublNauchMaterialFact',
    EGEVip='PublNauchMaterialVip',
    EGEFactVipolnenia='PublNauchMaterialFactVipolnenia',
    df=result_df,
    scaleEgeVip=scalePublNauchMaterialVip,
    scaleEgeNapr = scalePublNauchMaterialNapr,
    EGEFactBallCompare='PublNauchMaterialFactBallCompare',
    EGEFactFinal='PublNauchMaterialFactFinal',
    sred_C02=dictSred['Ц17']
)
result_df['PublNauchMaterialNapr'] = result_df['PublNauchMaterialFact'] / dictSred['Ц17']
result_df[['kafedra', 'PublNauchMaterialFact', 'PublNauchMaterialVip', 'PublNauchMaterialNapr','PublNauchMaterialFactVipolnenia', 'PublNauchMaterialFactBallCompare', 'PublNauchMaterialFactFinal']]


# In[161]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'PublNauchMaterialFact', 'PublNauchMaterialVip', 'PublNauchMaterialNapr','PublNauchMaterialFactVipolnenia', 'PublNauchMaterialFactBallCompare', 'PublNauchMaterialFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Подготовленные к публикации.xlsx', 'Подготовленные к публикации', scalePublNauchMaterialVip, scalePublNauchMaterialNapr)


# In[ ]:





# In[164]:


# Для Подготовка энциклопедии/справочника
result_df = process_data(
    EGEFact='PodgotovkaEnciklFact',
    EGEVip='PodgotovkaEnciklVip',
    EGEFactVipolnenia='PodgotovkaEnciklFactVipolnenia',
    df=result_df,
    scaleEgeVip=scalePodgotovkaEnciklVip,
    scaleEgeNapr = scalePodgotovkaEnciklNapr,
    EGEFactBallCompare='PodgotovkaEnciklFactBallCompare',
    EGEFactFinal='PodgotovkaEnciklFactFinal',
    sred_C02=dictSred['Ц18']
)
result_df['PodgotovkaEnciklNapr'] = result_df['PodgotovkaEnciklFact'] / dictSred['Ц18']
result_df[['kafedra', 'PodgotovkaEnciklFact', 'PodgotovkaEnciklVip', 'PodgotovkaEnciklNapr','PodgotovkaEnciklFactVipolnenia', 'PodgotovkaEnciklFactBallCompare', 'PodgotovkaEnciklFactFinal']]


# In[ ]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'PodgotovkaEnciklFact', 'PodgotovkaEnciklVip', 'PodgotovkaEnciklNapr','PodgotovkaEnciklFactVipolnenia', 'PodgotovkaEnciklFactBallCompare', 'PodgotovkaEnciklFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Подготовка энциклопедии.xlsx', 'Подготовка энциклопедии', scalePodgotovkaEnciklVip, scalePodgotovkaEnciklNapr)


# In[167]:


# Для Руководство работами студентов и аспирантов
result_df = process_data(
    EGEFact='RukovRabotyFact',
    EGEVip='RukovRabotyVip',
    EGEFactVipolnenia='RukovRabotyFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleRukovRabotyVip,
    scaleEgeNapr = scaleRukovRabotyNapr,
    EGEFactBallCompare='RukovRabotyFactBallCompare',
    EGEFactFinal='RukovRabotyFactFinal',
    sred_C02=dictSred['Ц20']
)
result_df['RukovRabotyNapr'] = result_df['RukovRabotyFact'] / dictSred['Ц20']
result_df[['kafedra', 'RukovRabotyFact', 'RukovRabotyVip', 'RukovRabotyNapr', 'RukovRabotyFactVipolnenia', 'RukovRabotyFactBallCompare', 'RukovRabotyFactFinal']]


# In[169]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'RukovRabotyFact', 'RukovRabotyVip', 'RukovRabotyNapr', 'RukovRabotyFactVipolnenia', 'RukovRabotyFactBallCompare', 'RukovRabotyFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Руководство работами студ.xlsx', 'Руководство работами', scaleRukovRabotyVip, scaleRukovRabotyNapr)



# In[171]:


# Для Инновации
result_df = process_data(
    EGEFact='InnovaciiFact',
    EGEVip='InnovaciiVip',
    EGEFactVipolnenia='InnovaciiFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleInnovaciiVip,
    scaleEgeNapr = scaleInnovaciiNapr,
    EGEFactBallCompare='InnovaciiFactBallCompare',
    EGEFactFinal='InnovaciiFactFinal',
    sred_C02=dictSred['Ц24']
)
result_df['InnovaciiNapr'] = result_df['InnovaciiFact'] / dictSred['Ц24']
result_df[['kafedra', 'InnovaciiFact', 'InnovaciiVip', 'InnovaciiNapr', 'InnovaciiFactVipolnenia', 'InnovaciiFactBallCompare', 'InnovaciiFactFinal']]


# In[173]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'InnovaciiFact', 'InnovaciiVip', 'InnovaciiNapr', 'InnovaciiFactVipolnenia', 'InnovaciiFactBallCompare', 'InnovaciiFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Инновации.xlsx', 'Инновации', scaleInnovaciiVip, scaleInnovaciiNapr)


# In[175]:


# Для Доля ППС, СЗП которых >= 200% от региональной
result_df = process_data(
    EGEFact='DolPPS200Fact',
    EGEVip='DolPPS200Vip',
    EGEFactVipolnenia='DolPPS200FactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleDolPPS200Vip,
    scaleEgeNapr = scaleDolPPS200Napr,
    EGEFactBallCompare='DolPPS200FactBallCompare',
    EGEFactFinal='DolPPS200FactFinal',
    sred_C02=dictSred['Ц25']
)
result_df['DolPPS200Napr'] = result_df['DolPPS200Fact'] / dictSred['Ц25']
result_df[['kafedra', 'DolPPS200Fact', 'DolPPS200Vip', 'DolPPS200Napr', 'DolPPS200FactVipolnenia', 'DolPPS200FactBallCompare', 'DolPPS200FactFinal']]


# In[177]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'DolPPS200Fact', 'DolPPS200Vip', 'DolPPS200Napr', 'DolPPS200FactVipolnenia', 'DolPPS200FactBallCompare', 'DolPPS200FactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Доля ППС, СЗП.xlsx', 'Доля ППС', scaleDolPPS200Vip, scaleDolPPS200Napr)


# In[179]:


# Для Доля общей численности НПР до 39 лет
result_df = process_data(
    EGEFact='DolMolNPRFact',
    EGEVip='DolMolNPRVip',
    EGEFactVipolnenia='DolMolNPRFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleDolMolNPRVip,
    scaleEgeNapr = scaleDolMolNPRNapr,
    EGEFactBallCompare='DolMolNPRFactBallCompare',
    EGEFactFinal='DolMolNPRFactFinal',
    sred_C02=dictSred['Ц26']
)
result_df['DolMolNPRNapr'] = result_df['DolMolNPRFact'] / dictSred['Ц26']
result_df[['kafedra', 'DolMolNPRFact', 'DolMolNPRVip', 'DolMolNPRNapr','DolMolNPRFactVipolnenia', 'DolMolNPRFactBallCompare', 'DolMolNPRFactFinal']]


# In[181]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'DolMolNPRFact', 'DolMolNPRVip', 'DolMolNPRNapr','DolMolNPRFactVipolnenia', 'DolMolNPRFactBallCompare', 'DolMolNPRFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'Доля общей численности НПР.xlsx', 'Доля общей численности', scaleDolMolNPRVip, scaleDolMolNPRNapr)


# In[ ]:





# In[184]:


# Для РИД
result_df = process_data(
    EGEFact='RIDFact',
    EGEVip='RIDVip',
    EGEFactVipolnenia='RIDFactVipolnenia',
    df=result_df,
    scaleEgeVip=scaleRidVip,
    scaleEgeNapr = scaleRidNapr,
    EGEFactBallCompare='RIDFactBallCompare',
    EGEFactFinal='RIDFactFinal',
    sred_C02=dictSred['Ц27']
)
result_df['RIDNapr'] = result_df['RIDFact'] / dictSred['Ц27']
result_df[['kafedra', 'RIDFact', 'RIDVip', 'RIDNapr', 'RIDFactVipolnenia', 'RIDFactBallCompare', 'RIDFactFinal']]


# In[186]:


# Список столбцов для сохранения
columns_to_save = ['kafedra', 'RIDFact', 'RIDVip', 'RIDNapr', 'RIDFactVipolnenia', 'RIDFactBallCompare', 'RIDFactFinal']
save_df_with_scales_to_excel(result_df, columns_to_save, 'РИД.xlsx', 'РИД', scaleRidVip, scaleRidNapr)


# In[ ]:





# In[189]:


print(result_df.columns.tolist())


# In[191]:


# Сохраняем DataFrame в CSV файл
result_df.to_csv('result.csv', index=False)


# In[193]:


# Очищаем DataFrame
result_dfNew = pd.DataFrame()

# Загружаем данные из CSV файла
result_dfNew = pd.read_csv('result.csv')


# In[195]:


#EGE + Trudoustroistvo + RID + PublNauchMaterial


# In[197]:


result_dfNew[['kafedra', 'EGEFactFinal','TrudoustroistvoFactFinal', 'RIDFactFinal','PublNauchMaterialFactFinal']]


# In[199]:


# Указываем столбцы для расчета
columns_to_average = ['EGEFactFinal', 'TrudoustroistvoFactFinal', 'RIDFactFinal', 'PublNauchMaterialFactFinal']

# Создаем новый столбец 'krit6'
result_dfNew['krit6'] = result_dfNew.apply(calculate_krit6, columns=columns_to_average, axis=1)


# In[201]:


result_dfNew[['kafedra', 'EGEFactFinal','TrudoustroistvoFactFinal', 'RIDFactFinal','PublNauchMaterialFactFinal','krit6']]


# In[203]:


print(result_dfNew.columns.tolist())


# In[205]:


result_dfNew[['kafedra', 'EGEFactFinal','TrudoustroistvoFactFinal', 'RIDFactFinal','PublNauchMaterialFactFinal','krit6']]


# In[207]:


print(list(result_dfNew.columns))


# In[209]:


columns_to_select = [
    'kafedra',    
    'AkadMobFactFinal',            # Академическая мобильность студентов и аспирантов, чел. (Ц13)
    'DolInostrFactFinal',          # Доля иностранных студентов (Ц11)
    'DolMagFactFinal',             # Доля магистров и аспирантов в общем контингенте, % (Ц01)
    'DolOstepenennykhFactFinal',   # Доля остепененных ППС, % (Ц06)
    'DolPPS200FactFinal',          # Доля ППС, СЗП (Ц25)
    'DolStorMagFactFinal',         # Доля сторонних магистров и аспирантов, % (Ц03)
    'DolCelevikovFactFinal',       # Доля целевиков, % (Ц04)
    'ZashchityFactFinal',          # Доля защит аспирантов в срок, % (Ц10)
    'DolMolNPRFactFinal',          # Доля ППС до 39 лет, % (Ц26)
    'ZayavAktivFactFinal',         # Заявочная активность, ед. (Ц16)
    'InnovaciiFactFinal',          # Инновации, ед. (Ц24)
    'KolInostrPrepodFactFinal',    # Количество иностранных преподавателей, чел. (Ц12)
    'KolPublikFactFinal',          # Количество публикаций в WoS (Ц08)
    'ObemNIOKRFactFinal',          # Объем НИОКР на ставку НПР, тыс. руб. (Ц09)
    'ObemPOUFactFinal'             # Объем ПОУ, тыс. руб. (Ц05)
]

# Вывод выбранных столбцов
selected_columns_df = result_dfNew[columns_to_select]


# In[211]:


selected_columns_df


# In[213]:


#Критерий 7


# In[215]:


columns_for_krit7 = [
    'AkadMobFactFinal',            # Академическая мобильность студентов и аспирантов, чел. (Ц13)
    'DolInostrFactFinal',          # Доля иностранных студентов (Ц11)
    'DolMagFactFinal',             # Доля магистров и аспирантов в общем контингенте, % (Ц01)
    'DolOstepenennykhFactFinal',   # Доля остепененных ППС, % (Ц06)
    'DolPPS200FactFinal',          # Доля ППС, СЗП (Ц25)
    'DolStorMagFactFinal',         # Доля сторонних магистров и аспирантов, % (Ц03)
    'DolCelevikovFactFinal',       # Доля целевиков, % (Ц04)
    'ZashchityFactFinal',          # Доля защит аспирантов в срок, % (Ц10)
    'DolMolNPRFactFinal',           # Доля ППС до 39 лет, % (Ц26)
    'ZayavAktivFactFinal',         # Заявочная активность, ед. (Ц16)
    'InnovaciiFactFinal',          # Инновации, ед. (Ц24)
    'KolInostrPrepodFactFinal',    # Количество иностранных преподавателей, чел. (Ц12)
    'KolPublikFactFinal',          # Количество публикаций в WoS (Ц08)
    'ObemNIOKRFactFinal',          # Объем НИОКР на ставку НПР, тыс. руб. (Ц09)
    'ObemPOUFactFinal'             # Объем ПОУ на ставку НПР, тыс. руб Ц05
]

# Функция для расчета krit7
def calculate_krit7(row, columns):
    # Извлекаем ненулевые значения из указанных столбцов
    non_zero_values = [value for value in row[columns] if value > 0]
    
    # Если ненулевых значений нет, возвращаем 0
    if not non_zero_values:
        return 0
    
    # Вычисляем среднее ненулевых значений
    avg_value = sum(non_zero_values) / len(non_zero_values)
    
    # Округляем до ближайшего кратного 5
    return round(avg_value / 5) * 5

# Применяем функцию к DataFrame
result_dfNew['krit7'] = result_dfNew.apply(calculate_krit7, columns=columns_for_krit7, axis=1)


# In[217]:


result_dfNew[['kafedra'] + columns_for_krit7 + ['krit7']]


# In[219]:


print(list(result_dfNew.columns))


# In[221]:


zz = list(result_dfNew.columns)


# In[223]:


type(zz)


# In[225]:


zz2 = ['kafedra', 
'DolMagFact', 'DolMagVip', 'DolMagNapr', 'DolMagFactVipolnenia', 'DolMagFactBallCompare', 'DolMagFactFinal', 
'EGEFact', 'EGEVip', 'EGENapr', 'EGEFactVipolnenia', 'EGEFactBallCompare', 'EGEFactFinal', 
'DolStorMagFact', 'DolStorMagVip', 'DolStorMagNapr', 'DolStorMagFactVipolnenia', 'DolStorMagFactBallCompare', 'DolStorMagFactFinal', 
'DolCelevikovFact', 'DolCelevikovVip', 'DolCelevikovNapr', 'DolCelevikovFactVipolnenia', 'DolCelevikovFactBallCompare', 'DolCelevikovFactFinal', 
'ObemPOUFact', 'ObemPOUVip', 'ObemPOUNapr', 'ObemPOUFactVipolnenia', 'ObemPOUFactBallCompare', 'ObemPOUFactFinal', 
'DolOstepenennykhFact', 'DolOstepenennykhVip', 'DolOstepenennykhNapr', 'DolOstepenennykhFactVipolnenia', 'DolOstepenennykhFactBallCompare', 'DolOstepenennykhFactFinal', 
'TrudoustroistvoFact', 'TrudoustroistvoVip', 'TrudoustroistvoNapr', 'TrudoustroistvoFactVipolnenia', 'TrudoustroistvoFactBallCompare', 'TrudoustroistvoFactFinal', 
'KolPublikFact', 'KolPublikVip', 'KolPublikNapr', 'KolPublikFactVipolnenia', 'KolPublikFactBallCompare', 'KolPublikFactFinal', 
'ObemNIOKRFact', 'ObemNIOKRVip', 'ObemNIOKRNapr', 'ObemNIOKRFactVipolnenia', 'ObemNIOKRFactBallCompare', 'ObemNIOKRFactFinal', 
'ZashchityFact', 'ZashchityVip', 'ZashchityNapr', 'ZashchityFactVipolnenia', 'ZashchityFactBallCompare', 'ZashchityFactFinal', 
'DolInostrFact', 'DolInostrVip', 'DolInostrNapr', 'DolInostrFactVipolnenia', 'DolInostrFactBallCompare', 'DolInostrFactFinal', 
'KolInostrPrepodFact', 'KolInostrPrepodVip', 'KolInostrPrepodNapr', 'KolInostrPrepodFactVipolnenia', 'KolInostrPrepodFactBallCompare', 'KolInostrPrepodFactFinal', 
'AkadMobFact', 'AkadMobVip', 'AkadMobNapr', 'AkadMobFactVipolnenia', 'AkadMobFactBallCompare', 'AkadMobFactFinal', 
'ZayavAktivFact', 'ZayavAktivVip', 'ZayavAktivNapr', 'ZayavAktivFactVipolnenia', 'ZayavAktivFactBallCompare', 'ZayavAktivFactFinal', 
'PublNauchMaterialFact', 'PublNauchMaterialVip', 'PublNauchMaterialNapr', 'PublNauchMaterialFactVipolnenia', 'PublNauchMaterialFactBallCompare', 'PublNauchMaterialFactFinal', 
'InnovaciiFact', 'InnovaciiVip', 'InnovaciiNapr', 'InnovaciiFactVipolnenia', 'InnovaciiFactBallCompare', 'InnovaciiFactFinal', 
'DolPPS200Fact', 'DolPPS200Vip', 'DolPPS200Napr', 'DolPPS200FactVipolnenia', 'DolPPS200FactBallCompare', 'DolPPS200FactFinal', 
'DolMolNPRFact', 'DolMolNPRVip', 'DolMolNPRNapr', 'DolMolNPRFactVipolnenia', 'DolMolNPRFactBallCompare', 'DolMolNPRFactFinal', 
'RIDFact', 'RIDVip', 'RIDNapr', 'RIDFactVipolnenia', 'RIDFactBallCompare', 'RIDFactFinal', 
'krit6', 'krit7', 
'PodgotovkaEnciklFact', 'PodgotovkaEnciklVip', 
'RukovRabotyFact', 'RukovRabotyVip', 
'UchastieVystavkiFact', 'UchastieVystavkiVip', 
'ChlenstvoSovetovFact', 'ChlenstvoSovetovVip', 
'ChlenstvoKomitetovFact', 'ChlenstvoKomitetovVip'
]


# In[227]:


result_dfNew2 = result_dfNew[zz2]


# In[229]:


type(zz2)


# In[231]:


# Находим элементы, которые есть в одном списке, но отсутствуют в другом
differences = set(zz).symmetric_difference(set(zz2))

# Выводим различия
print("Элементы, которые есть только в одном из списков:", differences)


# In[233]:


# Сохраняем DataFrame в CSV файл
result_dfNew2.to_csv('resultFinal.csv', index=False)


# In[235]:


# Specify the filename and sheet name
output_file = "output3.xlsx"
sheet_name = "Sheet1"

# Save the DataFrame to an Excel file with UTF-8 encoding
result_dfNew2.to_excel(output_file, index=False, sheet_name=sheet_name, engine="xlsxwriter")

print(f"DataFrame successfully saved to {output_file}")


# In[237]:


# Загрузка существующего файла Excel
input_file = "output4.xlsx"
sheet_name = "Sheet1"

column_mapping = {
    "kafedra": "Кафедра",
    "DolMagFact": "Факт доля магистров",
    "DolMagVip": "План доля магистров",
    "DolMagNapr": "Напряженность доля магистров",
    "DolMagFactVipolnenia": "Выполнение доля магистров",
    "DolMagFactBallCompare": "Сравнение баллов доля магистров",
    "DolMagFactFinal": "Итог доля магистров",
    "EGEFact": "Факт средний балл ЕГЭ",
    "EGEVip": "План средний балл ЕГЭ",
    "EGENapr": "Напряженность средний балл ЕГЭ",
    "EGEFactVipolnenia": "Выполнение средний балл ЕГЭ",
    "EGEFactBallCompare": "Сравнение баллов средний балл ЕГЭ",
    "EGEFactFinal": "Итог средний балл ЕГЭ",
    "DolStorMagFact": "Факт доля сторонних магистров",
    "DolStorMagVip": "План доля сторонних магистров",
    "DolStorMagNapr": "Напряженность доля сторонних магистров",
    "DolStorMagFactVipolnenia": "Выполнение доля сторонних магистров",
    "DolStorMagFactBallCompare": "Сравнение баллов доля сторонних магистров",
    "DolStorMagFactFinal": "Итог доля сторонних магистров",
    "DolCelevikovFact": "Факт доля целевиков",
    "DolCelevikovVip": "План доля целевиков",
    "DolCelevikovNapr": "Напряженность доля целевиков",
    "DolCelevikovFactVipolnenia": "Выполнение доля целевиков",
    "DolCelevikovFactBallCompare": "Сравнение баллов доля целевиков",
    "DolCelevikovFactFinal": "Итог доля целевиков",
    "ObemPOUFact": "Факт объем ПОУ",
    "ObemPOUVip": "План объем ПОУ",
    "ObemPOUNapr": "Напряженность объем ПОУ",
    "ObemPOUFactVipolnenia": "Выполнение объем ПОУ",
    "ObemPOUFactBallCompare": "Сравнение баллов объем ПОУ",
    "ObemPOUFactFinal": "Итог объем ПОУ",
    "DolOstepenennykhFact": "Факт доля остепененных ППС",
    "DolOstepenennykhVip": "План доля остепененных ППС",
    "DolOstepenennykhNapr": "Напряженность доля остепененных ППС",
    "DolOstepenennykhFactVipolnenia": "Выполнение доля остепененных ППС",
    "DolOstepenennykhFactBallCompare": "Сравнение баллов доля остепененных ППС",
    "DolOstepenennykhFactFinal": "Итог доля остепененных ППС",
    "TrudoustroistvoFact": "Факт трудоустройства",
    "TrudoustroistvoVip": "План трудоустройства",
    "TrudoustroistvoNapr": "Напряженность трудоустройства",
    "TrudoustroistvoFactVipolnenia": "Выполнение трудоустройства",
    "TrudoustroistvoFactBallCompare": "Сравнение баллов трудоустройства",
    "TrudoustroistvoFactFinal": "Итог трудоустройства",
    "KolPublikFact": "Факт публикаций",
    "KolPublikVip": "План публикаций",
    "KolPublikNapr": "Напряженность публикаций",
    "KolPublikFactVipolnenia": "Выполнение публикаций",
    "KolPublikFactBallCompare": "Сравнение баллов публикаций",
    "KolPublikFactFinal": "Итог публикаций",
    "ObemNIOKRFact": "Факт объем НИОКР",
    "ObemNIOKRVip": "План объем НИОКР",
    "ObemNIOKRNapr": "Напряженность объем НИОКР",
    "ObemNIOKRFactVipolnenia": "Выполнение объем НИОКР",
    "ObemNIOKRFactBallCompare": "Сравнение баллов объем НИОКР",
    "ObemNIOKRFactFinal": "Итог объем НИОКР",
    "ZashchityFact": "Факт защит аспирантов",
    "ZashchityVip": "План защит аспирантов",
    "ZashchityNapr": "Напряженность защит аспирантов",
    "ZashchityFactVipolnenia": "Выполнение защит аспирантов",
    "ZashchityFactBallCompare": "Сравнение баллов защит аспирантов",
    "ZashchityFactFinal": "Итог защит аспирантов",
    "DolInostrFact": "Факт доля иностранных студентов",
    "DolInostrVip": "План доля иностранных студентов",
    "DolInostrNapr": "Напряженность доля иностранных студентов",
    "DolInostrFactVipolnenia": "Выполнение доля иностранных студентов",
    "DolInostrFactBallCompare": "Сравнение баллов доля иностранных студентов",
    "DolInostrFactFinal": "Итог доля иностранных студентов",
    "KolInostrPrepodFact": "Факт иностранных преподавателей",
    "KolInostrPrepodVip": "План иностранных преподавателей",
    "KolInostrPrepodNapr": "Напряженность иностранных преподавателей",
    "KolInostrPrepodFactVipolnenia": "Выполнение иностранных преподавателей",
    "KolInostrPrepodFactBallCompare": "Сравнение баллов иностранных преподавателей",
    "KolInostrPrepodFactFinal": "Итог иностранных преподавателей",
    "AkadMobFact": "Факт академической мобильности",
    "AkadMobVip": "План академической мобильности",
    "AkadMobNapr": "Напряженность академической мобильности",
    "AkadMobFactVipolnenia": "Выполнение академической мобильности",
    "AkadMobFactBallCompare": "Сравнение баллов академической мобильности",
    "AkadMobFactFinal": "Итог академической мобильности",
    "ZayavAktivFact": "Факт заявочной активности",
    "ZayavAktivVip": "План заявочной активности",
    "ZayavAktivNapr": "Напряженность заявочной активности",
    "ZayavAktivFactVipolnenia": "Выполнение заявочной активности",
    "ZayavAktivFactBallCompare": "Сравнение баллов заявочной активности",
    "ZayavAktivFactFinal": "Итог заявочной активности",
    "PublNauchMaterialFact": "Факт научных материалов",
    "PublNauchMaterialVip": "План научных материалов",
    "PublNauchMaterialNapr": "Напряженность научных материалов",
    "PublNauchMaterialFactVipolnenia": "Выполнение научных материалов",
    "PublNauchMaterialFactBallCompare": "Сравнение баллов научных материалов",
    "PublNauchMaterialFactFinal": "Итог научных материалов",
    "InnovaciiFact": "Факт инноваций",
    "InnovaciiVip": "План инноваций",
    "InnovaciiNapr": "Напряженность инноваций",
    "InnovaciiFactVipolnenia": "Выполнение инноваций",
    "InnovaciiFactBallCompare": "Сравнение баллов инноваций",
    "InnovaciiFactFinal": "Итог инноваций",
    "DolPPS200Fact": "Факт доли ППС 200%",
    "DolPPS200Vip": "План доли ППС 200%",
    "DolPPS200Napr": "Напряженность доли ППС 200%",
    "DolPPS200FactVipolnenia": "Выполнение доли ППС 200%",
    "DolPPS200FactBallCompare": "Сравнение баллов доли ППС 200%",
    "DolPPS200FactFinal": "Итог доли ППС 200%",
    "DolMolNPRFact": "Факт доли молодых НПР",
    "DolMolNPRVip": "План доли молодых НПР",
    "DolMolNPRNapr": "Напряженность доли молодых НПР",
    "DolMolNPRFactVipolnenia": "Выполнение доли молодых НПР",
    "DolMolNPRFactBallCompare": "Сравнение баллов доли молодых НПР",
    "DolMolNPRFactFinal": "Итог доли молодых НПР",
    "RIDFact": "Факт РИД",
    "RIDVip": "План РИД",
    "RIDNapr": "Напряженность РИД",
    "RIDFactVipolnenia": "Выполнение РИД",
    "RIDFactBallCompare": "Сравнение баллов РИД",
    "RIDFactFinal": "Итог РИД",
    "krit6": "Критерий 6",
    "krit7": "Критерий 7",
    "PodgotovkaEnciklFact": "Факт подготовки энциклопедии",
    "PodgotovkaEnciklVip": "План подготовки энциклопедии",
    "RukovRabotyFact": "Факт руководства работами",
    "RukovRabotyVip": "План руководства работами",
    "UchastieVystavkiFact": "Факт участия в выставках",
    "UchastieVystavkiVip": "План участия в выставках",
    "ChlenstvoSovetovFact": "Факт членства в советах",
    "ChlenstvoSovetovVip": "План членства в советах",
    "ChlenstvoKomitetovFact": "Факт членства в комитетах",
    "ChlenstvoKomitetovVip": "План членства в комитетах"
}

# Переименование заголовков
result_dfNew2.rename(columns=column_mapping, inplace=True)

# Сохранение измененного файла Excel
output_file = "output_updated.xlsx"
result_dfNew2.to_excel(output_file, index=False, sheet_name=sheet_name, engine="xlsxwriter")

print(f"Заголовки успешно обновлены и файл сохранен в {output_file}")


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:


# Сохраняем DataFrame в CSV файл
result_dfNew.to_csv('resultFinal.csv', index=False)


# In[ ]:


get_ipython().system('pip install pandas openpyxl xlsxwriter')


# In[ ]:


# Specify the filename and sheet name
output_file = "output2.xlsx"
sheet_name = "Sheet1"

# Save the DataFrame to an Excel file with UTF-8 encoding
result_dfNew.to_excel(output_file, index=False, sheet_name=sheet_name, engine="xlsxwriter")

print(f"DataFrame successfully saved to {output_file}")


# In[ ]:


pwd


# In[ ]:


# Очищаем DataFrame
df2 = pd.DataFrame()

# Загружаем данные из CSV файла
df2 = pd.read_csv('dannie.csv')


# In[ ]:


# Specify the filename and sheet name
output_file = "dannie.xlsx"
sheet_name = "Sheet1"

# Save the DataFrame to an Excel file with UTF-8 encoding
df2.to_excel(output_file, index=False, sheet_name=sheet_name, engine="xlsxwriter")

print(f"DataFrame successfully saved to {output_file}")


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:


result = get_sredBallFinal_values(df, 'kafedra', 'sredBallVipolnenia', 'sredBallCompare', 'sredBallFinal')
print("\n".join(result))


# In[ ]:




